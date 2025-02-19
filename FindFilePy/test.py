import os
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook

def main():
    # 如果不想弹窗选择文件夹，可将此部分注释并直接指定 folder_path
    root = tk.Tk()
    root.withdraw()  # 隐藏主窗口
    folder_path = filedialog.askdirectory(title="请选择包含 Excel 文件的文件夹")
    
    # 如果想手动写死文件夹路径，则可直接指定，如：
    # folder_path = r"D:\测试目录"

    if not folder_path:
        print("未选择文件夹，程序退出。")
        return
    
    # 目标 sheet 名列表（小写）
    target_sheets = ["post", "put", "delete", "get"]

    # 遍历文件夹下所有 Excel 文件
    for file_name in os.listdir(folder_path):
        # 判断是否为 Excel 文件（可按需扩展）
        if file_name.endswith(".xlsx") or file_name.endswith(".xls"):
            file_path = os.path.join(folder_path, file_name)
            try:
                wb = load_workbook(file_path)
            except Exception as e:
                print(f"打开 Excel 文件失败：{file_name}，错误：{e}")
                continue

            print(f"正在处理文件：{file_name} ...")
            
            # 遍历所有工作表
            for sheet_name in wb.sheetnames:
                # 只处理目标 sheet
                if sheet_name.lower() in target_sheets:
                    ws = wb[sheet_name]
                    print(f"  >> Sheet 名：{sheet_name}")

                    # 假设我们只取 A, B, C 三列；如果需要更多列可自行调整
                    # 从第一行开始遍历，直到最后一行
                    data_list = []
                    for row in ws.iter_rows(min_row=1, values_only=True):
                        # row 是一个元组，如 (a_val, b_val, c_val, ...)
                        # 这里先取前 3 列
                        a_val = row[0] if len(row) > 0 else None
                        b_val = row[1] if len(row) > 1 else None
                        c_val = row[2] if len(row) > 2 else None

                        # 根据自己的逻辑，这里可能需要判断是否为空
                        # 或者对 a_val, b_val, c_val 做进一步分析
                        data_list.append((a_val, b_val, c_val))
                    
                    # 根据 data_list 中的数据做需要的整理/展示
                    # 这里只是简单打印
                    for item in data_list:
                        print("      A列: {}, B列: {}, C列: {}".format(
                            item[0], item[1], item[2]
                        ))
            
            print(f"文件 {file_name} 处理完毕。\n")
    
    print("所有文件处理结束。")

if __name__ == "__main__":
    main()