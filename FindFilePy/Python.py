import os
import re
import openpyxl
from openpyxl import Workbook

import tkinter as tk
from tkinter import filedialog

#######################
# 去除非法字符的正则
#######################
ILLEGAL_CHARACTERS_RE = re.compile(
    r'[\x00-\x08\x0B-\x0C\x0E-\x1F\x7F-\x9F'  # 常见的ASCII控制字符
    r'\uD800-\uDFFF\uFFFE\uFFFF]'            # 代理区和特殊不可用字符
)

def remove_illegal_characters(value):
    """
    去除字符串中 openpyxl 无法处理的非法字符（控制符、代理区等）。
    """
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value

#######################
# 解析 .cma 文件
#######################
def parse_cma_file(file_path):
    """
    解析单个 .cma 文件，返回所有 *LINE 块。
    每个块会以(块内容字符串)的形式返回，方便后续再做键值对提取。
    """
    results = []
    # 使用Shift_JIS编码来读取文件，如果遇到不认识的字符就替换
    with open(file_path, 'r', encoding='shift_jis', errors='replace') as f:
        lines = f.readlines()
    
    # 去掉首尾空格和换行
    lines = [line.strip() for line in lines]
    
    current_block = []
    collecting_block = False

    for line in lines:
        if line.startswith("*LINE"):
            # 如果之前有正在收集的块未结束，则先结束它
            if collecting_block and current_block:
                combined_line = " ".join(current_block)
                results.append(combined_line)
                current_block = []
            
            collecting_block = True
            current_block.append(line)
        else:
            if collecting_block:
                current_block.append(line)
                if ";" in line:
                    # 如果本行包含分号，说明块结束
                    combined_line = " ".join(current_block)
                    results.append(combined_line)
                    collecting_block = False
                    current_block = []
    
    # 文件末尾如果还有没结束的块，也存一下
    if collecting_block and current_block:
        combined_line = " ".join(current_block)
        results.append(combined_line)
    
    return results

#######################
# 从 *LINE 文本块中提取键值对
#######################
def extract_key_values(block_str):
    """
    给定块的字符串，比如 "*LINE A=xxx B=yyy ;"
    提取所有 "key=value" 形式，返回一个 dict，例如 {"A": "xxx", "B": "yyy"}
    """
    # 1) 去掉 *LINE
    block_str = block_str.replace("*LINE", "")
    # 2) 按分号拆分，只保留分号前部分
    block_str = block_str.split(";")[0]
    block_str = block_str.strip()
    
    # 用正则找出所有 k=v
    pattern = r'(\w+)\s*=\s*([^=\s]+)'
    pairs = re.findall(pattern, block_str)
    
    kv_dict = {}
    for k, v in pairs:
        kv_dict[k] = v
    return kv_dict

#######################
# 遍历文件夹收集所有数据
#######################
def scan_cma_and_collect(folder_path):
    """
    递归遍历 folder_path 下所有 .cma 文件，解析其中的所有块，收集键值对。
    返回 (all_records, all_keys):
      - all_records: 列表，每个元素形如 {"__filename": 文件名, "A": "xxx", "B": "yyy", ...}
      - all_keys: set，包含除 __filename 以外所有出现过的键
    """
    all_records = []
    all_keys = set()
    
    for root, _, files in os.walk(folder_path):
        for f in files:
            if f.lower().endswith(".cma"):
                full_path = os.path.join(root, f)
                blocks = parse_cma_file(full_path)
                
                # 对每个 *LINE block 提取 key-value
                for block in blocks:
                    kv_dict = extract_key_values(block)
                    if kv_dict:
                        kv_dict["__filename"] = f
                        all_records.append(kv_dict)
                        for k in kv_dict.keys():
                            if k != "__filename":
                                all_keys.add(k)
    
    return all_records, all_keys

#######################
# 写入 Excel
#######################
def write_to_excel(records, keys, output_excel="output.xlsx"):
    """
    将数据写入Excel：
      - 第一行： 文件名 | (所有出现过的key按字母表排序)
      - 之后每行对应一个块的数据
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "CMA数据"

    # 对收集到的键做排序，保证列顺序一致
    sorted_keys = sorted(keys)

    # 写表头
    ws.cell(row=1, column=1, value="文件名")
    col_index = 2
    for k in sorted_keys:
        ws.cell(row=1, column=col_index, value=remove_illegal_characters(k))
        col_index += 1

    # 写每条记录
    current_row = 2
    for rec in records:
        # 第一列写文件名
        file_name = rec["__filename"]
        ws.cell(row=current_row, column=1, value=remove_illegal_characters(file_name))

        # 后续列写各键值
        col_index = 2
        for k in sorted_keys:
            val = rec.get(k, "")
            ws.cell(row=current_row, column=col_index, value=remove_illegal_characters(val))
            col_index += 1

        current_row += 1

    wb.save(output_excel)
    print(f"数据已写入 {output_excel} 文件。")

#######################
# 主入口
#######################
if __name__ == "__main__":
    # 使用 tkinter 打开文件夹选择对话框（Windows可弹出资源管理器）
    root = tk.Tk()
    root.withdraw()

    folder_to_search = filedialog.askdirectory(title='请选择要扫描的文件夹')
    if folder_to_search:
        all_records, all_keys = scan_cma_and_collect(folder_to_search)
        
        # 写入 Excel
        output_file = "cma_data.xlsx"
        write_to_excel(all_records, all_keys, output_file)
    else:
        print("未选择任何文件夹，程序退出。")